# -*- coding: utf-8 -*-
"""
Created on Tue Oct 31 02:53:52 2017

@author: rishi
"""

from tkinter import filedialog
from tkinter import *
import openpyxl

from openpyxl import load_workbook

#Define Window Geometry

main = Tk()
main.geometry("1024x768")
main.title("Window Title")

#Define Empty Cells to be Filled in by Excel File & Calculation

def OpenDataInputSpreadsheetCallBack():
    main.iconify()
    file_path = filedialog.askopenfilename(initialdir = "file_path_goes_here",title = "Choose Input Spreadsheet",filetypes = (("Excel 2010 files","*.xlsx"),("Excel 2003 Files","*.xls")))
    wb = load_workbook(filename = file_path, read_only=True)
    
    class Textbox(object):
        text = None
    ws = wb.active
    series_of_textboxes = [Textbox(),Textbox(),Textbox(),Textbox()]
# start reading from row 2
    for i, row in enumerate( ws.iter_rows(min_row=2) ):
        series_of_textboxes[i].text = ' '.join(cell.value for cell in row)

    print( series_of_textboxes[0].text )  
    for row in ws.iter_rows():
        for cell in row:
            if (cell.value==None):
                pass
            else:
                print(cell.value)
#
#
#Function to display empty rows and columns              
#
height = 5
width = 6
for x in range(1,height+1): #Rows
   for y in range(width): #Columns
        b = Entry(main, text='')
        b.grid(row=x, column=y)    
#      
# Define Buttons    

b1 = Button(main, text = "Open Data Input Spreadsheet", command = OpenDataInputSpreadsheetCallBack)
b1.place(x = 1,y = 120)
b2 = Button(main, text='Quit', command=main.destroy)
b2.place(x = 1,y = 150)
##
##
### Initialize Column Headers
Label(main, text="Header1").grid(row=0, column=0, sticky=W)
Label(main, text="Header2").grid(row=0, column=1, sticky=W)
Label(main, text="Header3").grid(row=0, column=2, sticky=W)
Label(main, text="Header4").grid(row=0, column=3, sticky=W)
Label(main, text="Header5").grid(row=0, column=4, sticky=W)
Label(main, text="Header6").grid(row=0, column=5, sticky=W)







###
# Define a function to close the window.
def quit(event=None):
    main.destroy()
# Cause pressing <Esc> to close the window.
main.bind('<Escape>', quit)
#
#
main.mainloop()